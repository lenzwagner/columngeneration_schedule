import os
import pandas as pd
import openpyxl

# Sets
work = pd.read_excel(r'./data/Arzt.xlsx', sheet_name='Arzt')
df = pd.read_excel(r'./data/NF.xlsx', sheet_name='NF')
df1 = pd.read_excel(r'./data/NF.xlsx', sheet_name='Shift')
I = work['Id'].tolist()
W_I = work['Weekend'].tolist()
T = df['Day'].tolist()
K = df1['Shift'].tolist()
S_T = df1['Hours'].tolist()
I_T = work['WT'].tolist()

Min_WD_i = [2, 3, 2, 3]
Max_WD_i = [5, 5, 5, 6]
S_T = {a: c for a, c in zip(K, S_T)}
I_T = {a: d for a, d in zip(I, I_T)}
W_I = {a: e for a, e in zip(I, W_I)}

Min_WD_i = {a: f for a, f in zip(I, Min_WD_i)}
Max_WD_i = {a: g for a, g in zip(I, Max_WD_i)}
print(W_I)

I_list1 = pd.DataFrame(I, columns=['I'])
T_list1 = pd.DataFrame(T, columns=['T'])
K_list1 = pd.DataFrame(K, columns=['K'])
DataDF = pd.concat([I_list1, T_list1, K_list1], axis=1)

Demand_Dict = {}
workbook = openpyxl.load_workbook(r'./data/NF.xlsx')
worksheet = workbook['NF']
for row in worksheet.iter_rows(min_row=2, values_only=True):
    for i in range(1, len(row)):
        cell_value = row[i]
        if isinstance(cell_value, str) and cell_value.startswith('='):
            cell = worksheet.cell(row=row[0], column=i)
            cell_value = cell.value
        if isinstance(cell_value, int):
            Demand_Dict[(int(row[0]), i)] = cell_value
workbook.close()
